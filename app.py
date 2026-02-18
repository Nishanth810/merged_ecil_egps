from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException, TimeoutException, NoSuchWindowException
import time
import pandas as pd
import smtplib
from email.message import EmailMessage
import os
import json
import re
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

print("Starting COMBINED ECIL + EGPS (ISRO) tender extraction...")

# ==============================
# BROWSER WINDOW HELPERS
# ==============================
def safe_maximize_window(driver):
    try:
        driver.maximize_window()
    except Exception as e:
        print(f"! Maximize window skipped: {e}")
        try:
            driver.set_window_size(1400, 900)
        except Exception as size_error:
            print(f"! Window resize skipped: {size_error}")

def safe_close_extra_windows(driver, main_window):
    try:
        all_windows = driver.window_handles
        for window in all_windows:
            if window != main_window:
                try:
                    driver.switch_to.window(window)
                    driver.close()
                except:
                    pass
        driver.switch_to.window(main_window)
    except Exception as e:
        print(f"    ⚠️ Window cleanup error: {e}")
        try:
            driver.switch_to.window(main_window)
        except:
            pass

def get_available_filename(base_name):
    if not os.path.exists(base_name):
        return base_name, False
    try:
        with open(base_name, 'a'):
            pass
        return base_name, False
    except PermissionError:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        name_parts = base_name.rsplit('.', 1)
        if len(name_parts) == 2:
            new_name = f"{name_parts[0]}_{timestamp}.{name_parts[1]}"
        else:
            new_name = f"{base_name}_{timestamp}"
        print(f"⚠️ Original file is open/locked: {base_name}")
        print(f"✓ Using timestamped filename: {new_name}")
        return new_name, True

# ==============================
# LOAD ENVIRONMENT VARIABLES
# ==============================
load_dotenv()

sender_email = os.getenv("SENDER_EMAIL")
app_password = os.getenv("APP_PASSWORD")
receiver_emails_str = os.getenv("RECEIVER_EMAILS", "")
receiver_emails = [email.strip() for email in receiver_emails_str.split(",") if email.strip()]

if not sender_email or not app_password:
    print("⚠️ WARNING: Email credentials not found in .env file")
if not receiver_emails:
    print("⚠️ WARNING: No receiver emails found in .env file")

base_file_name = "combined_tenders.xlsx"
HISTORY_FILE = "combined_tender_history.json"

# ==============================
# TENDER HISTORY MANAGEMENT
# ==============================
def load_tender_history():
    if Path(HISTORY_FILE).exists():
        try:
            with open(HISTORY_FILE, "r") as f:
                return json.load(f)
        except Exception as e:
            print(f"⚠️ Error loading history: {e}")
            return {"ECIL": {}, "EGPS": {}}
    return {"ECIL": {}, "EGPS": {}}

def save_tender_history(history):
    try:
        with open(HISTORY_FILE, "w") as f:
            json.dump(history, f, indent=2)
    except Exception as e:
        print(f"⚠️ Error saving history: {e}")

def check_date_changes(current_data, history, source):
    changes = []
    for entry in current_data:
        tender_no = entry[0]
        description = entry[2]
        closing_date = entry[3]  # index 3 = closing date
        if tender_no in history[source]:
            old_closing = history[source][tender_no].get("closing_date", "")
            if old_closing and closing_date and old_closing != closing_date:
                changes.append({
                    "tender_no": tender_no,
                    "old_date": old_closing,
                    "new_date": closing_date,
                    "source": source,
                    "description": description
                })
    return changes

def update_tender_history(ecil_data, egps_data):
    history = load_tender_history()
    # ecil entry: [tender_no, centre, description, closing_date, published_date, opening_date, link, doc_links]
    for entry in ecil_data:
        tender_no = entry[0]
        history["ECIL"][tender_no] = {
            "description": entry[2],
            "closing_date": entry[3],
            "last_seen": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
    # egps entry: [tender_no, centre, description, closing_date, published_date, opening_date, link, doc_links]
    for entry in egps_data:
        tender_no = entry[0]
        history["EGPS"][tender_no] = {
            "centre": entry[1],
            "description": entry[2],
            "closing_date": entry[3],
            "last_seen": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
    save_tender_history(history)
    print(f"✅ History updated: ECIL={len(ecil_data)}, EGPS={len(egps_data)}")

# ==============================
# ECIL DOCUMENT EXTRACTION
# ==============================
def extract_ecil_documents(driver):
    doc_links = []
    try:
        time.sleep(2)
        try:
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
        except TimeoutException:
            print("    ⚠️ Page load timeout")
            return []

        pdf_elements = driver.find_elements(By.XPATH,
            "//a[contains(@href, '.pdf') or contains(@href, '.PDF')]"
        )
        print(f"    Found {len(pdf_elements)} PDF links")

        for pdf in pdf_elements:
            try:
                url = pdf.get_attribute("href")
                name = pdf.text.strip()
                if not name or name in ["--NA--", "Download", "View"]:
                    name = url.split("/")[-1].replace('.pdf', '').replace('.PDF', '')
                if url and "--NA--" not in name and url not in [x[1] for x in doc_links]:
                    doc_links.append((name, url))
                    print(f"    ✓ {name[:60]}")
            except Exception as e:
                print(f"    ⚠️ Error extracting PDF link: {e}")
                continue

        print(f"    📎 Total: {len(doc_links)}")
        return doc_links
    except Exception as e:
        print(f"    ✗ Error: {e}")
        return []

# ==============================
# EGPS (ISRO) DOCUMENT EXTRACTION + PUBLISHED DATE
# ==============================
def extract_egps_documents_and_published_date(driver, page_type="View"):
    """
    Extract document links AND published date from EGPS View page.
    The Tender Schedule section contains:
      - Published Date
      - Bid Submission Start Date
      - Bid Submission End Date
      - Bid Opening Date
    Returns: (doc_links, published_date)
    """
    doc_links = []
    published_date = ""

    try:
        time.sleep(2)
        try:
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
        except TimeoutException:
            print("    ⚠️ Page load timeout")
            return [], ""

        # ── Extract Published Date from Tender Schedule ──
        try:
            # Look for the label "Published Date" and grab the adjacent value
            pub_date_elements = driver.find_elements(By.XPATH,
                "//*[contains(text(), 'Published Date')]"
            )
            for elem in pub_date_elements:
                try:
                    # Try sibling or parent-sibling approach
                    parent = elem.find_element(By.XPATH, "..")
                    parent_text = parent.text.strip()
                    # Pattern: "Published Date : 16-02-2026 17:20 IST"
                    match = re.search(
                        r'Published Date\s*[:\-]?\s*(\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{2,4}(?:\s+\d{1,2}:\d{2}(?::\d{2})?(?:\s*[APap][Mm])?(?:\s*IST)?)?)',
                        parent_text, re.IGNORECASE
                    )
                    if match:
                        published_date = match.group(1).strip()
                        print(f"    📅 Published Date: {published_date}")
                        break

                    # Try next sibling td or div
                    try:
                        sibling = elem.find_element(By.XPATH, "following-sibling::*[1]")
                        sib_text = sibling.text.strip()
                        if sib_text and re.search(r'\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{2,4}', sib_text):
                            published_date = sib_text
                            print(f"    📅 Published Date (sibling): {published_date}")
                            break
                    except:
                        pass
                except:
                    continue

            # Fallback: search full page text
            if not published_date:
                body_text = driver.find_element(By.TAG_NAME, "body").text
                match = re.search(
                    r'Published Date\s*[:\-]?\s*(\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{2,4}(?:\s+\d{1,2}:\d{2}(?::\d{2})?(?:\s*[APap][Mm])?(?:\s*IST)?)?)',
                    body_text, re.IGNORECASE
                )
                if match:
                    published_date = match.group(1).strip()
                    print(f"    📅 Published Date (body): {published_date}")

        except Exception as e:
            print(f"    ⚠️ Published date extraction error: {e}")

        # ── Extract Document Links ──
        all_links = driver.find_elements(By.XPATH,
            "//a[contains(@href, '.pdf') or contains(@href, '.PDF') or "
            "contains(@href, 'viewDocument') or contains(@href, 'downloadDocument') or "
            "contains(@onclick, 'viewDocument') or contains(@data-url, 'viewDocument')]"
        )
        print(f"    Found {len(all_links)} potential links")

        for link in all_links:
            try:
                href = link.get_attribute("href")
                data_url = link.get_attribute("data-url")
                onclick = link.get_attribute("onclick")
                text = link.text.strip()

                url = href
                if data_url and not url:
                    url = "https://eproc.isro.gov.in" + data_url
                elif onclick and not url:
                    match = re.search(r"'([^']*(?:viewDocument|downloadDocument)[^']*)'", onclick)
                    if match:
                        url = "https://eproc.isro.gov.in" + match.group(1)

                if url and (
                    '.pdf' in url.lower() or
                    'viewDocument' in url or
                    'downloadDocument' in url
                ):
                    if not text or text in ["View", "Download", "Open", "Click Here"]:
                        try:
                            parent_row = link.find_element(By.XPATH, "./ancestor::tr[1]")
                            cells = parent_row.find_elements(By.TAG_NAME, "td")
                            for cell in cells[:3]:
                                cell_text = cell.text.strip()
                                if cell_text and cell_text not in ["View", "Download", "Open", ""]:
                                    text = cell_text
                                    break
                        except:
                            pass
                        if not text or text in ["View", "Download", "Open"]:
                            text = url.split('/')[-1].replace('.pdf', '').replace('.PDF', '') or "Document"

                    if page_type == "Corrigendum" and not text.startswith("Corrigendum"):
                        text = f"Corrigendum - {text}"
                    elif page_type == "View" and not text.startswith("View"):
                        text = f"View - {text}"

                    if url not in [x[1] for x in doc_links]:
                        doc_links.append((text, url))
                        print(f"    ✓ {text[:60]}")
            except Exception as e:
                print(f"    ⚠️ Error extracting link: {e}")
                continue

        print(f"    📎 Total docs: {len(doc_links)}")
        return doc_links, published_date

    except Exception as e:
        print(f"    ✗ Error: {e}")
        return [], ""

# ==============================
# SCRAPE ECIL TENDERS
# ==============================
def scrape_ecil():
    print("\n" + "="*60)
    print("SCRAPING ECIL TENDERS")
    print("="*60)

    options = webdriver.ChromeOptions()
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-gpu')

    driver = None
    try:
        driver = webdriver.Chrome(options=options)
        wait = WebDriverWait(driver, 15)

        print("Loading ECIL website...")
        driver.get("https://etenders.ecil.co.in/")
        safe_maximize_window(driver)
        time.sleep(3)

        try:
            nit_button = wait.until(
                EC.element_to_be_clickable((By.XPATH, "//*[contains(text(),'Active Public NIT')]"))
            )
            nit_button.click()
            print("✓ Clicked Active Public NIT")
        except Exception as e:
            print(f"✗ Error clicking NIT button: {e}")
            if driver:
                driver.quit()
            return []

        time.sleep(3)

        data = []

        # Detect total pages
        total_pages = 1
        try:
            pagination_links = driver.find_elements(By.XPATH,
                "//a[@href and text()[string-length(normalize-space(.)) > 0]]"
            )
            page_numbers = []
            for link in pagination_links:
                text = link.text.strip()
                if text.isdigit() and 1 <= len(text) <= 2:
                    try:
                        page_num = int(text)
                        if 1 <= page_num <= 100:
                            page_numbers.append(page_num)
                    except ValueError:
                        continue
            if page_numbers:
                total_pages = max(page_numbers)
                print(f"✓ Pagination detected: {sorted(set(page_numbers))}")

            if total_pages == 1:
                page_text_elements = driver.find_elements(By.XPATH,
                    "//*[contains(text(), 'of') or contains(@class, 'page')]"
                )
                for elem in page_text_elements:
                    text = elem.text.strip()
                    match = re.search(r'of\s+(\d+)', text, re.IGNORECASE)
                    if match:
                        total_pages = int(match.group(1))
                        print(f"✓ Found page count in text: {text}")
                        break
        except Exception as e:
            print(f"⚠️ Pagination detection error: {e}")
            total_pages = 1

        print(f"Total pages detected: {total_pages}\n")

        for page in range(1, total_pages + 1):
            print(f"\nPAGE {page}/{total_pages}")
            time.sleep(2)

            rows = driver.find_elements(By.XPATH,
                "//table//tbody//tr[td] | "
                "//table[contains(@class, 'table')]//tr[td]"
            )
            print(f"  Found {len(rows)} rows to process")

            row_index = 0
            rows_processed_this_page = 0

            while row_index < len(rows):
                try:
                    row = rows[row_index]
                    cols = row.find_elements(By.TAG_NAME, "td")
                except StaleElementReferenceException:
                    print("  ⚠️ Stale element, refreshing rows...")
                    rows = driver.find_elements(By.XPATH,
                        "//table//tbody//tr[td] | "
                        "//table[contains(@class, 'table')]//tr[td]"
                    )
                    continue

                # ECIL table columns (from screenshot):
                # 0=Section Unit, 1=Tender No, 2=Published Date, 3=Description, 4=Type, 5=Due Date/Time, 6=Due Days
                if len(cols) >= 6:
                    tender_no = cols[1].text.strip()

                    if (
                        not tender_no or
                        tender_no == "" or
                        "NIT" in tender_no or
                        "Section" in tender_no or
                        "Tender" in tender_no or
                        tender_no.lower() in ["no.", "number", "tender no."] or
                        (tender_no.isdigit() and len(tender_no) <= 2)
                    ):
                        row_index += 1
                        continue

                    if len(tender_no) < 5:
                        row_index += 1
                        continue

                    # ── Extract Published Date from col[2] ──
                    published_date = cols[2].text.strip() if len(cols) > 2 else ""

                    description = cols[3].text.strip() if len(cols) > 3 else ""
                    due_date = cols[5].text.strip() if len(cols) > 5 else ""

                    if not description or len(description) < 5:
                        row_index += 1
                        continue

                    try:
                        tender_link = cols[1].find_element(By.TAG_NAME, "a").get_attribute("href")
                    except:
                        tender_link = ""

                    print(f"📋 {tender_no} | Published: {published_date}")
                    rows_processed_this_page += 1

                    doc_links = []
                    if tender_link:
                        print(f"  🔍 Extracting documents...")
                        main_window = driver.current_window_handle
                        try:
                            driver.execute_script(f"window.open('{tender_link}');")
                            time.sleep(2)
                            if len(driver.window_handles) > 1:
                                driver.switch_to.window(driver.window_handles[-1])
                                doc_links = extract_ecil_documents(driver)
                                driver.close()
                                driver.switch_to.window(main_window)
                            else:
                                print("  ⚠️ New window didn't open")
                        except Exception as e:
                            print(f"  ✗ Document extraction error: {e}")
                            safe_close_extra_windows(driver, main_window)

                    print(f"  ✅ {len(doc_links)} docs")

                    # Entry structure:
                    # [tender_no, centre, description, closing_date, published_date, opening_date, link, doc_links]
                    data.append([
                        tender_no,          # 0 tender_no
                        "-----",            # 1 centre (ECIL doesn't have centre)
                        description,        # 2 description
                        due_date,           # 3 closing/due date
                        published_date,     # 4 published date  ← NEW
                        "-----",            # 5 opening date (ECIL doesn't show separately)
                        tender_link,        # 6 link
                        doc_links           # 7 docs
                    ])

                row_index += 1

            print(f"  ✓ Processed {rows_processed_this_page} tenders on this page")

            if page < total_pages:
                try:
                    next_page_link = driver.find_element(By.XPATH,
                        f"//a[@href and normalize-space(text())='{page+1}']"
                    )
                    print(f"  🔄 Navigating to page {page+1}...")
                    driver.execute_script("arguments[0].scrollIntoView(true);", next_page_link)
                    time.sleep(0.5)
                    driver.execute_script("arguments[0].click();", next_page_link)
                    time.sleep(3)
                    time.sleep(2)
                except Exception as e:
                    print(f"  ⚠️ Could not navigate to page {page+1}: {e}")
                    break

        print(f"\n✅ ECIL: {len(data)} tenders scraped")
        return data

    except Exception as e:
        print(f"\n❌ ECIL scraping failed: {e}")
        import traceback
        traceback.print_exc()
        return []

    finally:
        if driver:
            try:
                driver.quit()
            except:
                pass

# ==============================
# SCRAPE EGPS (ISRO) TENDERS
# ==============================
def scrape_egps():
    print("\n" + "="*60)
    print("SCRAPING EGPS (ISRO) TENDERS")
    print("="*60)

    options = webdriver.ChromeOptions()
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-gpu')

    driver = None
    try:
        driver = webdriver.Chrome(options=options)
        wait = WebDriverWait(driver, 15)

        print("Loading EGPS website...")
        driver.get("https://eproc.isro.gov.in/home.html")
        safe_maximize_window(driver)
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "table")))
        time.sleep(2)

        data = []

        page_elements = driver.find_elements(By.XPATH, "//a[text()[number(.)=number(.)]]")
        pages = [int(el.text.strip()) for el in page_elements if el.text.strip().isdigit()]
        total_pages = max(pages) if pages else 1

        print(f"Total pages: {total_pages}\n")

        for page in range(1, total_pages + 1):
            print(f"\nPAGE {page}/{total_pages}")
            time.sleep(2)

            rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")

            for row in rows:
                cols = row.find_elements(By.TAG_NAME, "td")
                if len(cols) < 6:
                    continue

                tender_no = cols[0].text.strip()
                centre = cols[1].text.strip()
                description = cols[2].text.strip()
                closing = cols[3].text.strip()
                opening = cols[4].text.strip()

                print(f"📋 {tender_no}")

                doc_links = []
                published_date = ""

                action_links = cols[5].find_elements(By.TAG_NAME, "a")

                tender_pdf = None
                view_link = None
                corrigendum_link = None

                for l in action_links:
                    text = l.text.strip()
                    href = l.get_attribute("href") or ""
                    data_url = l.get_attribute("data-url") or ""

                    if "Tender Document" in text and href:
                        tender_pdf = href
                    if "homeTenderView" in data_url:
                        view_link = "https://eproc.isro.gov.in" + data_url
                    if "viewCorrigendum" in data_url or "Corrigendum" in text:
                        if data_url:
                            corrigendum_link = "https://eproc.isro.gov.in" + data_url
                        else:
                            corrigendum_link = href

                if tender_pdf:
                    doc_links.append(("Tender Document", tender_pdf))
                    print(f"  📄 Main PDF")

                main_window = driver.current_window_handle

                # ── Open View page: extract docs AND published date ──
                if view_link:
                    print(f"  🔍 View page (docs + published date)...")
                    try:
                        driver.execute_script("window.open(arguments[0]);", view_link)
                        time.sleep(2)

                        if len(driver.window_handles) > 1:
                            driver.switch_to.window(driver.window_handles[-1])
                            view_docs, pub_date = extract_egps_documents_and_published_date(driver, "View")

                            if pub_date:
                                published_date = pub_date

                            for name, url in view_docs:
                                if url not in [x[1] for x in doc_links]:
                                    doc_links.append((name, url))

                            driver.close()
                            driver.switch_to.window(main_window)

                    except Exception as e:
                        print(f"  ✗ View error: {e}")
                        safe_close_extra_windows(driver, main_window)

                if corrigendum_link:
                    print(f"  📝 Corrigendum docs...")
                    try:
                        driver.execute_script("window.open(arguments[0]);", corrigendum_link)
                        time.sleep(2)

                        if len(driver.window_handles) > 1:
                            driver.switch_to.window(driver.window_handles[-1])
                            corr_docs, _ = extract_egps_documents_and_published_date(driver, "Corrigendum")

                            for name, url in corr_docs:
                                if url not in [x[1] for x in doc_links]:
                                    doc_links.append((name, url))

                            driver.close()
                            driver.switch_to.window(main_window)

                    except Exception as e:
                        print(f"  ✗ Corrigendum error: {e}")
                        safe_close_extra_windows(driver, main_window)

                print(f"  ✅ {len(doc_links)} docs | Published: {published_date}")

                # Entry structure:
                # [tender_no, centre, description, closing_date, published_date, opening_date, link, doc_links]
                data.append([
                    tender_no,          # 0 tender_no
                    centre,             # 1 centre
                    description,        # 2 description
                    closing,            # 3 closing date
                    published_date,     # 4 published date  ← NEW
                    opening,            # 5 opening date
                    view_link or "",    # 6 link
                    doc_links           # 7 docs
                ])

            if page < total_pages:
                try:
                    next_page = driver.find_element(By.XPATH, f"//a[text()='{page+1}']")
                    driver.execute_script("arguments[0].click();", next_page)
                    time.sleep(2)
                except Exception as e:
                    print(f"  ⚠️ Could not navigate to page {page+1}: {e}")
                    break

        print(f"\n✅ EGPS: {len(data)} tenders scraped")
        return data

    except Exception as e:
        print(f"\n❌ EGPS scraping failed: {e}")
        import traceback
        traceback.print_exc()
        return []

    finally:
        if driver:
            try:
                driver.quit()
            except:
                pass

# ==============================
# MAIN EXECUTION
# ==============================

try:
    ecil_data = scrape_ecil()
    egps_data = scrape_egps()

    print(f"\n{'='*60}")
    print("COMBINING DATA")
    print(f"{'='*60}")

    max_docs_ecil = max([len(entry[7]) for entry in ecil_data], default=0)
    max_docs_egps = max([len(entry[7]) for entry in egps_data], default=0)
    max_docs = max(max_docs_ecil, max_docs_egps)

    print(f"Max docs ECIL: {max_docs_ecil}")
    print(f"Max docs EGPS: {max_docs_egps}")
    print(f"Max docs overall: {max_docs}")

    history = load_tender_history()

    ecil_changes = check_date_changes(ecil_data, history, "ECIL")
    egps_changes = check_date_changes(egps_data, history, "EGPS")

    print(f"\n⚠️ Date Changes Detected:")
    print(f"  ECIL - {len(ecil_changes)} closing date changes")
    print(f"  EGPS - {len(egps_changes)} closing date changes")

    if ecil_changes:
        print("\n  ECIL Changes:")
        for change in ecil_changes:
            print(f"    {change['tender_no']}: {change['old_date']} → {change['new_date']}")

    if egps_changes:
        print("\n  EGPS Changes:")
        for change in egps_changes:
            print(f"    {change['tender_no']}: {change['old_date']} → {change['new_date']}")

    changed_tenders = set()
    for change in ecil_changes + egps_changes:
        changed_tenders.add((change['source'], change['tender_no']))

    ecil_new = []
    ecil_changed = []
    ecil_existing = []
    egps_new = []
    egps_changed = []
    egps_existing = []

    for entry in ecil_data:
        tender_no = entry[0]
        if tender_no not in history["ECIL"]:
            ecil_new.append(["NEW", "ECIL"] + entry)
        elif ("ECIL", tender_no) in changed_tenders:
            ecil_changed.append(["DATE CHANGED", "ECIL"] + entry)
        else:
            ecil_existing.append(["EXISTING", "ECIL"] + entry)

    for entry in egps_data:
        tender_no = entry[0]
        if tender_no not in history["EGPS"]:
            egps_new.append(["NEW", "EGPS"] + entry)
        elif ("EGPS", tender_no) in changed_tenders:
            egps_changed.append(["DATE CHANGED", "EGPS"] + entry)
        else:
            egps_existing.append(["EXISTING", "EGPS"] + entry)

    print(f"\n📊 Categorization:")
    print(f"  ECIL - New: {len(ecil_new)}, Date Changed: {len(ecil_changed)}, Existing: {len(ecil_existing)}")
    print(f"  EGPS - New: {len(egps_new)}, Date Changed: {len(egps_changed)}, Existing: {len(egps_existing)}")

    all_tenders = (ecil_new + egps_new +
                   ecil_changed + egps_changed +
                   ecil_existing + egps_existing)

    print("\nBuilding Excel...")

    rows = []
    for entry in all_tenders:
        # entry = [status, source, tender_no, centre, description, closing_date, published_date, opening_date, link, docs]
        status       = entry[0]
        source       = entry[1]
        tender_no    = entry[2]
        centre       = entry[3]
        description  = entry[4]
        closing_date = entry[5]
        published_date = entry[6]
        opening_date = entry[7]
        link         = entry[8]
        docs         = entry[9]

        # ── Column order as requested ──
        # Status | Source | Tender Number | Published Date | Bid Opening Date | Bid Closing Date |
        # Centre/Organization | Description | Tender Link | Doc1 Name | Doc1 Link | ...
        row = [
            status,
            source,
            tender_no,
            published_date,
            opening_date,
            closing_date,
            centre,
            description,
            link
        ]

        for i in range(max_docs):
            if i < len(docs):
                row.append(docs[i][0])
                row.append(docs[i][1])
            else:
                row.append("")
                row.append("")

        rows.append(row)

    columns = [
        "Status",
        "Source",
        "Tender Number",
        "Published Date",
        "Bid Opening Date",
        "Bid Closing Date",
        "Centre/Organization",
        "Description",
        "Tender Link"
    ]

    for i in range(1, max_docs + 1):
        columns.append(f"Document {i} Name")
        columns.append(f"Document {i} Link")

    combined_df = pd.DataFrame(rows, columns=columns)

    file_name, is_timestamped = get_available_filename(base_file_name)

    with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
        combined_df.to_excel(writer, sheet_name="All Tenders", index=False)

        ecil_df = combined_df[combined_df["Source"] == "ECIL"]
        if not ecil_df.empty:
            ecil_df.to_excel(writer, sheet_name="ECIL", index=False)

        egps_df = combined_df[combined_df["Source"] == "EGPS"]
        if not egps_df.empty:
            egps_df.to_excel(writer, sheet_name="EGPS", index=False)

        if not egps_df.empty:
            centres = sorted(egps_df["Centre/Organization"].unique())
            for centre in centres:
                if centre and centre != "-----":
                    centre_df = egps_df[egps_df["Centre/Organization"] == centre]
                    sheet_name = centre[:31].replace('/', '-').replace('\\', '-').replace('*', '').replace('[', '').replace(']', '')
                    try:
                        centre_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    except Exception as e:
                        print(f"⚠️ Could not create sheet for {centre}: {e}")

    print("Formatting Excel...")

    wb = load_workbook(file_name)

    green_fill  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    gray_fill   = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        headers = {}

        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            headers[cell_value] = col
            ws.cell(row=1, column=col).fill = header_fill
            ws.cell(row=1, column=col).font = Font(bold=True, color="FFFFFF")

        for row in range(2, ws.max_row + 1):
            status_cell = ws.cell(row=row, column=headers.get("Status", 1))

            if status_cell.value == "NEW":
                row_fill = green_fill
            elif status_cell.value == "DATE CHANGED":
                row_fill = yellow_fill
            else:
                row_fill = gray_fill

            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = row_fill

            for i in range(1, max_docs + 1):
                name_col = headers.get(f"Document {i} Name")
                link_col = headers.get(f"Document {i} Link")

                if name_col and link_col:
                    name = ws.cell(row=row, column=name_col).value
                    url  = ws.cell(row=row, column=link_col).value

                    if name and url:
                        cell = ws.cell(row=row, column=name_col)
                        cell.hyperlink = url
                        cell.font = Font(color="0563C1", underline="single")

        # Hide raw link columns
        for i in range(1, max_docs + 1):
            link_col = headers.get(f"Document {i} Link")
            if link_col:
                col_letter = ws.cell(row=1, column=link_col).column_letter
                ws.column_dimensions[col_letter].hidden = True

        # Also hide Tender Link column (hyperlink is embedded in Tender Number cell or keep visible)
        # Auto-fit visible columns
        for col in range(1, min(ws.max_column + 1, 20)):
            column_letter = ws.cell(row=1, column=col).column_letter
            if not ws.column_dimensions[column_letter].hidden:
                max_length = 0
                for row_num in range(1, min(ws.max_row + 1, 100)):
                    try:
                        cell_value = str(ws.cell(row=row_num, column=col).value or "")
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

    # Apply hyperlinks to Tender Number column as well
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        headers = {}
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            headers[cell_value] = col

        tender_num_col = headers.get("Tender Number")
        tender_link_col = headers.get("Tender Link")

        if tender_num_col and tender_link_col:
            for row in range(2, ws.max_row + 1):
                tender_num_cell = ws.cell(row=row, column=tender_num_col)
                tender_link_val = ws.cell(row=row, column=tender_link_col).value
                if tender_num_cell.value and tender_link_val:
                    tender_num_cell.hyperlink = tender_link_val
                    current_font = tender_num_cell.font
                    tender_num_cell.font = Font(
                        color="0563C1", underline="single", bold=current_font.bold
                    )

    wb.save(file_name)

    print(f"✅ Excel created: {file_name}")
    print(f"   Sheets: {len(wb.sheetnames)}")

    if is_timestamped:
        print(f"\n⚠️ NOTE: Original file was locked, created: {file_name}")

    update_tender_history(ecil_data, egps_data)

    # ── Email ──
    if sender_email and app_password and receiver_emails:
        print("\nSending email...")

        msg = EmailMessage()
        msg["From"] = sender_email
        msg["To"] = ", ".join(receiver_emails)

        subject_parts = []
        if len(ecil_new) + len(egps_new) > 0:
            subject_parts.append(f"{len(ecil_new) + len(egps_new)} NEW")
        if len(ecil_changes) + len(egps_changes) > 0:
            subject_parts.append(f"{len(ecil_changes) + len(egps_changes)} DATE CHANGED")

        msg["Subject"] = f"Combined Tender Update - {' | '.join(subject_parts)}" if subject_parts else "Combined Tender Update - No Changes"

        change_details = ""
        if ecil_changes or egps_changes:
            change_details = "\n\n⚠️ CLOSING DATE CHANGES:\n"
            change_details += "="*60 + "\n"

            if ecil_changes:
                change_details += "\n📌 ECIL Tenders:\n"
                for change in ecil_changes:
                    change_details += f"  • {change['tender_no']}\n"
                    change_details += f"    Old Date: {change['old_date']}\n"
                    change_details += f"    New Date: {change['new_date']}\n"
                    change_details += f"    Description: {change['description'][:80]}...\n\n"

            if egps_changes:
                change_details += "\n📌 EGPS (ISRO) Tenders:\n"
                for change in egps_changes:
                    change_details += f"  • {change['tender_no']}\n"
                    change_details += f"    Old Date: {change['old_date']}\n"
                    change_details += f"    New Date: {change['new_date']}\n"
                    change_details += f"    Description: {change['description'][:80]}...\n\n"

        email_body = f"""Combined ECIL + EGPS (ISRO) Tender Report
{datetime.now().strftime("%Y-%m-%d %I:%M %p")}

📊 ECIL Summary:
  🆕 NEW: {len(ecil_new)}
  ⚠️ DATE CHANGED: {len(ecil_changed)}
  📋 EXISTING: {len(ecil_existing)}
  📝 Total: {len(ecil_data)}

📊 EGPS (ISRO) Summary:
  🆕 NEW: {len(egps_new)}
  ⚠️ DATE CHANGED: {len(egps_changed)}
  📋 EXISTING: {len(egps_existing)}
  📝 Total: {len(egps_data)}

📊 Overall:
  🆕 Total NEW: {len(ecil_new) + len(egps_new)}
  ⚠️ Total DATE CHANGED: {len(ecil_changed) + len(egps_changed)}
  📋 Total EXISTING: {len(ecil_existing) + len(egps_existing)}
  📝 Grand Total: {len(all_tenders)}
  📄 Max Documents/Tender: {max_docs}
  📂 Total Sheets: {len(wb.sheetnames)}
{change_details}
⚠️ IMPORTANT: Check DATE CHANGED tenders - closing dates have been modified!
"""
        msg.set_content(email_body)

        with open(file_name, "rb") as f:
            msg.add_attachment(
                f.read(),
                maintype="application",
                subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=file_name
            )

        try:
            with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
                smtp.login(sender_email, app_password)
                smtp.send_message(msg)
            print("✅ Email sent!")
        except Exception as e:
            print(f"❌ Email sending failed: {e}")
    else:
        print("\n⚠️ Email not sent - credentials not configured")

    print(f"\n🎉 COMPLETED!")
    print(f"   ECIL: {len(ecil_new)} new | {len(ecil_changed)} date changed | {len(ecil_existing)} existing")
    print(f"   EGPS: {len(egps_new)} new | {len(egps_changed)} date changed | {len(egps_existing)} existing")
    print(f"   TOTAL: {len(ecil_new)+len(egps_new)} new | {len(ecil_changed)+len(egps_changed)} date changed | {len(ecil_existing)+len(egps_existing)} existing")

except Exception as e:
    print(f"\n❌ CRITICAL ERROR: {e}")
    import traceback
    traceback.print_exc()
